from __future__ import annotations
import os
import re
from typing import Dict, Any, List, Set

import pandas as pd  # pip install pandas openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

_INVALID_SHEET_CHARS = r'[:\\/*?\[\]]'  # Excel-invalid characters
_MAX_SHEET_LEN = 31

# Header style: solid green background + white bold font
_HEADER_FILL = PatternFill(fill_type="solid", start_color="FF2E7D32", end_color="FF2E7D32")  # #2E7D32
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _safe_sheet_name(raw_title: str, taken: Set[str]) -> str:
    """
    Create a safe Excel sheet name from a raw title.
    
    Ensures the sheet name is valid for Excel by removing invalid characters,
    handling length limits, and avoiding duplicates.

    :param raw_title: Original title to convert to sheet name
    :param taken: Set of already used sheet names to avoid conflicts
    :return: Safe Excel sheet name that doesn't conflict with existing names
    """
    name = (raw_title or "Untitled").strip()
    name = re.sub(_INVALID_SHEET_CHARS, "_", name)
    name = re.sub(r"\s+", " ", name)
    name = name[:_MAX_SHEET_LEN] if name else "Sheet"

    base = name or "Sheet"
    candidate = base
    i = 1
    while candidate in taken or not candidate:
        suffix = f"_{i}"
        candidate = (base[: _MAX_SHEET_LEN - len(suffix)] + suffix) if len(base) + len(suffix) > _MAX_SHEET_LEN else base + suffix
        i += 1
    taken.add(candidate)
    return candidate

def _style_header(ws, ncols: int) -> None:
    """
    Apply styling to the header row of an Excel worksheet.
    
    Styles the first row with green background, white bold font, and center alignment.
    Also freezes the panes below the header row.

    :param ws: OpenPyXL worksheet object to style
    :param ncols: Number of columns in the worksheet
    :return: None
    """
    # Style first row (header) and freeze panes below it
    ws.freeze_panes = "A2"
    for idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

def _autosize_columns(ws, df: pd.DataFrame) -> None:
    """
    Automatically size columns in an Excel worksheet based on content.
    
    Calculates optimal column widths based on header text and sample data
    from the first 200 rows for performance.

    :param ws: OpenPyXL worksheet object to resize
    :param df: Pandas DataFrame containing the data
    :return: None
    """
    # Basic autosize based on header + sample of values
    for i, col in enumerate(df.columns, start=1):
        header = str(col) if col is not None else ""
        max_len = len(header)
        # sample first ~200 rows for performance
        for val in df.iloc[:200, i - 1].astype(str).values:
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 60)

def write_structured_excel(excel_path: str, items: List[Dict[str, Any]]) -> str | None:
    """
    Write a list of structured data items into an Excel workbook.
    
    Each item becomes a separate worksheet with styled headers. The function
    handles sheet name sanitization, header styling, and column autosizing.

    :param excel_path: Path where the Excel file will be saved
    :param items: List of dictionaries, each containing:
                 - 'title': Sheet title (optional)
                 - 'headers': List of column headers (optional)
                 - 'rows': List of data rows (optional)
    :return: Path to the written Excel file if successful, None if no items provided
    """
    if not items:
        return None

    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)

    taken: Set[str] = set()
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        for item in items:
            title = item.get("title") or "Untitled"
            headers = item.get("headers") or []
            rows = item.get("rows") or []

            sheet_name = _safe_sheet_name(title, taken)

            # Build DataFrame; if headers mismatch row width, pandas will handle (pad/truncate)
            df = pd.DataFrame(rows, columns=headers if headers else None)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Style header + autosize
            ws = writer.sheets[sheet_name]
            _style_header(ws, ncols=df.shape[1] if df.shape[1] else 1)
            _autosize_columns(ws, df)

    return excel_path